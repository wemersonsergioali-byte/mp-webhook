import logging
import io
import re
import requests
from collections import defaultdict
from datetime import datetime
from flask import Flask, request, jsonify
from telegram import Update, Document, Bot
from telegram.ext import (
    ApplicationBuilder,
    MessageHandler,
    CommandHandler,
    ContextTypes,
    ConversationHandler,
    filters
)
from openpyxl import load_workbook, Workbook
import threading
import sqlite3

# ===================== CONFIGURAÇÃO =====================
TELEGRAM_TOKEN = "8505967080:AAGXzmGZYUK3BxrTEM0hd_sQhq6uZEEePa4"
MP_ACCESS_TOKEN = "APP_USR-264234346131232-071723-2b11d40f943d9721d869863410833122-777482543"

# ===================== LOG =====================
logging.basicConfig(
    format="%(asctime)s - %(levelname)s - %(message)s",
    level=logging.INFO
)

# ===================== FLASK =====================
app = Flask(__name__)

@app.route("/", methods=["GET"])
def home():
    return "Webhook Mercado Pago ONLINE", 200

@app.route("/webhook", methods=["POST"])
def webhook():
    try:
        data = request.get_json(silent=True)
        if not data:
            return jsonify({"status": "ok"}), 200

        payment_id = data.get("data", {}).get("id")
        if payment_id:
            payment = consultar_pagamento(payment_id)
            if payment.get("status") == "approved":
                telegram_id = payment.get("external_reference")
                creditos = int(payment.get("transaction_amount", 0))
                with get_db() as conn:
                    conn.execute("""
                        INSERT INTO usuarios (telegram_id, creditos)
                        VALUES (?, ?)
                        ON CONFLICT(telegram_id)
                        DO UPDATE SET creditos = creditos + ?
                    """, (telegram_id, creditos, creditos))
                    try:
                        bot = Bot(token=TELEGRAM_TOKEN)
                        bot.send_message(
                            chat_id=telegram_id,
                            text=f"✅ Pagamento aprovado! {creditos} créditos adicionados à sua conta."
                        )
                    except Exception as e:
                        print("Erro enviando mensagem Telegram:", e)
        return jsonify({"status": "ok"}), 200
    except Exception as e:
        print("Erro no webhook:", e)
        return jsonify({"status": "ok"}), 200

def consultar_pagamento(payment_id):
    url = f"https://api.mercadopago.com/v1/payments/{payment_id}"
    headers = {"Authorization": f"Bearer {MP_ACCESS_TOKEN}"}
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        return response.json()
    else:
        return {"error": True, "status_code": response.status_code, "response": response.text}

# ===================== BANCO DE DADOS =====================
def get_db():
    conn = sqlite3.connect("database.db", check_same_thread=False)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS usuarios (
            telegram_id INTEGER PRIMARY KEY,
            creditos INTEGER DEFAULT 1
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS pagamentos (
            payment_id TEXT PRIMARY KEY,
            telegram_id INTEGER,
            creditos INTEGER,
            status TEXT
        )
    """)
    return conn

# ===================== PLANOS =====================
ESCOLHENDO_PLANO = 1
PLANOS = {
    "1": {"creditos": 30, "valor": 19.90},
    "2": {"creditos": 60, "valor": 36.90},
    "3": {"creditos": 90, "valor": 51.90},
    "4": {"creditos": 120, "valor": 62.90},
}

# ===================== FUNÇÕES AUXILIARES =====================
def consultar_cep(cep):
    cep = str(cep).replace("-", "").strip()
    if not cep.isdigit() or len(cep) != 8:
        return None
    url = f"https://opencep.com/v1/{cep}"
    try:
        response = requests.get(url, timeout=5)
        if response.status_code == 200:
            data = response.json()
            return data.get("logradouro")
    except Exception as e:
        logging.warning(f"Erro consultando CEP {cep}: {e}")
    return None

def extrair_chave(endereco: str):
    endereco = endereco.strip()
    match = re.match(r"(.+?)(\d+)", endereco)
    if match:
        logradouro = match.group(1).strip().lower()
        numero = match.group(2)
        return f"{logradouro}|{numero}"
    return endereco.lower()

def corrigir_planilha(file_bytes):
    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    colunas_remover = ["AT ID", "Stop", "SPX TN"]
    for col in reversed(range(1, len(headers) + 1)):
        if headers[col - 1] in colunas_remover:
            ws.delete_cols(col)
    headers = [cell.value for cell in ws[1]]
    idx_seq = headers.index("Sequence") + 1
    idx_dest = headers.index("Destination Address") + 1
    idx_cep = headers.index("Zipcode/Postal code") + 1

    for row in range(2, ws.max_row + 1):
        cep = ws.cell(row=row, column=idx_cep).value
        endereco_original = str(ws.cell(row=row, column=idx_dest).value or "")
        logradouro_api = consultar_cep(cep)
        if logradouro_api:
            match = re.match(r"^[^\d]*", endereco_original.strip())
            if match:
                resto = endereco_original[len(match.group(0)):].strip(", ")
                novo_endereco = f"{logradouro_api}, {resto}" if resto else logradouro_api
                ws.cell(row=row, column=idx_dest, value=novo_endereco)

    agrupado = defaultdict(list)
    dados_extra = {}
    for row in range(2, ws.max_row + 1):
        seq = str(ws.cell(row=row, column=idx_seq).value or "").strip()
        endereco = str(ws.cell(row=row, column=idx_dest).value or "").strip()
        if endereco:
            chave = extrair_chave(endereco)
            agrupado[chave].append(seq)
            if chave not in dados_extra:
                dados_extra[chave] = [
                    ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)
                ]
                dados_extra[chave][headers.index("Destination Address")] = endereco

    wb_novo = Workbook()
    ws_novo = wb_novo.active
    ws_novo.append(headers)
    for chave, seqs in agrupado.items():
        linha = dados_extra[chave][:]
        total = len(seqs)
        seq_final = ",".join(seqs)
        if total > 1:
            seq_final = f"{seq_final}; Total: {total} pacotes"
        linha[headers.index("Sequence")] = seq_final
        ws_novo.append(linha)

    output = io.BytesIO()
    wb_novo.save(output)
    output.seek(0)
    return output

def saudacao():
    hora = datetime.now().hour
    if 5 <= hora < 12:
        return "Bom dia ☀️"
    elif 12 <= hora < 18:
        return "Boa tarde 🌤️"
    else:
        return "Boa noite 🌙"

# ===================== BOT TELEGRAM =====================
async def saldo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    telegram_id = update.effective_user.id
    with get_db() as conn:
        row = conn.execute("SELECT creditos FROM usuarios WHERE telegram_id=?", (telegram_id,)).fetchone()
    creditos = row[0] if row else 0
    await update.message.reply_text(f"💰 Seu saldo atual: {creditos} créditos")

async def comprar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    texto = "🌟 Ótimo! Escolha um plano:\n"
    for k, p in PLANOS.items():
        texto += f"{k}️⃣ {p['creditos']} créditos por R$ {p['valor']}\n"
    texto += "\nDigite o número da opção (1, 2, 3 ou 4)"
    await update.message.reply_text(texto)
    return ESCOLHENDO_PLANO

async def escolher_plano(update: Update, context: ContextTypes.DEFAULT_TYPE):
    escolha = update.message.text.strip()
    if escolha not in PLANOS:
        await update.message.reply_text("⚠️ Opção inválida. Digite 1,2,3 ou 4")
        return ESCOLHENDO_PLANO

    plano = PLANOS[escolha]
    creditos = plano["creditos"]
    valor = plano["valor"]

    headers = {"Authorization": f"Bearer {MP_ACCESS_TOKEN}", "Content-Type": "application/json"}
    payload = {
        "transaction_amount": valor,
        "description": f"Pacote de {creditos} créditos",
        "payment_method_id": "pix",
        "payer": {"email": "cliente@email.com"},
        "external_reference": str(update.effective_user.id)
    }

    try:
        response = requests.post("https://api.mercadopago.com/v1/payments", json=payload, headers=headers)
        data = response.json()
        payment_id = str(data.get("id"))

        with get_db() as conn:
            conn.execute("INSERT INTO pagamentos (payment_id, telegram_id, creditos, status) VALUES (?, ?, ?, ?)",
                         (payment_id, update.effective_user.id, creditos, "pending"))

        pi = data.get("point_of_interaction")
        if pi and "transaction_data" in pi and "qr_code" in pi["transaction_data"]:
            copia_cola = pi["transaction_data"]["qr_code"]
            await update.message.reply_text(f"💳 PIX gerado:\n`{copia_cola}`", parse_mode="Markdown")
        else:
            await update.message.reply_text(
                "⚠️ Pagamento registrado. O QR Code será enviado automaticamente via webhook quando confirmado."
            )
    except Exception as e:
        await update.message.reply_text(f"❌ Erro ao gerar PIX: {e}")

    return ConversationHandler.END

async def cancelar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Operação cancelada.")
    return ConversationHandler.END

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    document: Document = update.message.document
    user_name = update.message.from_user.first_name or "Usuário"
    logging.info(f"Recebida planilha de {user_name}")

    if document.mime_type != "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        await update.message.reply_text("⚠️ Envie apenas arquivos Excel (.xlsx)")
        return

    await update.message.reply_text(f"{saudacao()}, {user_name}! ⏳ Processando sua planilha...")
    file = await context.bot.get_file(document.file_id)
    file_bytes = await file.download_as_bytearray()
    planilha_corrigida = corrigir_planilha(file_bytes)

    hoje = datetime.now().strftime("%d-%m-%Y")
    nome_completo = (update.message.from_user.full_name or user_name).upper()
    filename = f"Rota-Atualizada-{hoje} {nome_completo}.xlsx"

    await update.message.reply_document(document=planilha_corrigida, filename=filename)
    await update.message.reply_text(f"✅ Planilha atualizada com sucesso, {user_name}!")

# ===================== START =====================
def start_telegram():
    app_telegram = ApplicationBuilder().token(TELEGRAM_TOKEN).build()
    app_telegram.add_handler(CommandHandler("saldo", saldo))
    comprar_handler = ConversationHandler(
        entry_points=[CommandHandler("comprar", comprar)],
        states={ESCOLHENDO_PLANO: [MessageHandler(filters.TEXT & ~filters.COMMAND, escolher_plano)]},
        fallbacks=[CommandHandler("cancelar", cancelar)]
    )
    app_telegram.add_handler(comprar_handler)
    app_telegram.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    logging.info("🤖 Bot Telegram iniciado e aguardando ações...")
    app_telegram.run_polling()


if __name__ == "__main__":
    # Flask em thread separada
    threading.Thread(target=lambda: app.run(host="0.0.0.0", port=8080), daemon=True).start()
    start_telegram()
